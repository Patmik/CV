'''program od wyceniania'''

import openpyxl as xl
import pyautogui
import pyperclip
import time
from selenium import webdriver
#import cv2
#from selenium.webdriver.common.keys import Keys

# klikanie az do zdków
def logowanie():
    x = pyautogui.locateOnScreen('zaloguj.png')
    y = pyautogui.center(x)
    pyautogui.click(y)
    time.sleep(6)
    
    a=pyautogui.locateOnScreen('otworz_baan.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(6)
    
    a=pyautogui.locateOnScreen('1_kontrakty.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(3)

    
    a=pyautogui.locateOnScreen('2_kontrakty.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(2)
    
    a=pyautogui.locateOnScreen('3_zlecenia.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(2)

    print('Zalogowano.')

#otwiera przegladarke i loguje do zd
def open_browser(): 
    #otwiera przeglądarke
    browser = webdriver.Ie()

    browser.get('http://plwaw1mm01.ad.agisfs.com:8312/webui/servlet/login?default')
    time.sleep(15)
    logowanie()
    print('Przeglądarka otwarta.')


# tworzy liste zd do wycen z excela
def process_workbook(filename):
    wb=xl.load_workbook(filename)
    sheet = wb['Arkusz1']
    

    for row in range (5,sheet.max_row+1):
        cell=sheet.cell(row,3)
        zd_to_list = cell.value
        zd_list.append(zd_to_list)
        print(f'{zd_to_list} pobrano')
        
    
    wb.save(filename)

    return zd_list
    


#wpisuje listy zd oraz ich statusow do nowego excela
def excel_with_results(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Arkusz1']
    i=4
    k=4

    for  zd_number in zd_list:

        cell = sheet.cell(i, 2)
        cell.value = str(zd_number)
        i=i+1

    for  zd_statu in zd_status:

        cell = sheet.cell(k, 3)
        cell.value = str(zd_statu)
        k=k+1


    wb.save(filename)

    print('Numery ZD zostały wycenione')

#szuka przycisku wyceniania i klika
def dolarek():
    #time.sleep(1)
    a=pyautogui.locateOnScreen('wycen.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(0.5)


#szuka przycisku wyceniania bladego i klika
def dolarek_z():
    time.sleep(1)
    a=pyautogui.locateOnScreen('wycenione.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(0.5)
    
#wkleja to co znajduje się w schowku i aceptuje (zd w tym przypadku)
def wklej(p):
    pyautogui.typewrite(p,0.1)
    time.sleep(0.5)
    pyautogui.press('enter')

#test funkcji przesuwania myszy    
def przesun():
    pyautogui.moveTo(1300,300,duration=0.25)
    pyautogui.click()


#przejezdza myszą na okno wpisywania zd i kasuje jego zawartosc
def kasuj():
    pyautogui.moveTo(481,265,duration=0.25)
    pyautogui.click()
    time.sleep(0.5)
    pyautogui.press('backspace')
    print('skasowano')
    time.sleep(1)


#lokalizuje okno do wpisywania zd
def okno():
    time.sleep(0.5)
    a=pyautogui.locateOnScreen('okienko_4_zd.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(0.5)


#zaznacza pasek zd do wyceny
def zaznacz():
    time.sleep(0.5)
    a=pyautogui.locateOnScreen('zaznaczenie.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(0.5)


#wychodzi z okna wpisywania zd
def klik():
    time.sleep(0.5)
    a=pyautogui.locateOnScreen('klik.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(0.5)

#sprawdza czy jest w pakiecie finanse_____________________________________________________________________________________________________________________________________________________
def sprawdzenie():

    if pyautogui.locateOnScreen('wycenione3.png'):
        
        zd_status.append("Już było wycenione")
        print('Już było wycenione')
        time.sleep(2)

    elif pyautogui.locateOnScreen('zamkniete.png'):
        
        zd_status.append("Status: zamkniete")
        print('Status: zamkniete')
        time.sleep(2)
        
    else:
        time.sleep(1)
        a=pyautogui.locateOnScreen('wejdz.png')
        b=pyautogui.center(a)
        pyautogui.click(b)
        time.sleep(2)

        time.sleep(1)
        a=pyautogui.locateOnScreen('inne.png')
        b=pyautogui.center(a)
        pyautogui.click(b)
        time.sleep(2)


        if pyautogui.locateOnScreen('finanse.png')== None:
            time.sleep(1.5)
            a=pyautogui.locateOnScreen('wyjdz.png')
            b=pyautogui.center(a)
            pyautogui.click(b)
            time.sleep(0.5)
            print(" Nie wystawiliśmy faktury")
            zd_status.append('Nie wystawiliśmy faktury')
        else:
            time.sleep(1)
            a=pyautogui.locateOnScreen('wyjdz.png')
            b=pyautogui.center(a)
            pyautogui.click(b)
            time.sleep(0.5)
            cost()

    
#except Exception:
       # print('nie ma finansów')
       # zd_status.append("Brak możliwości wyceny")
       # time.sleep(0.5)


    


#przypadki kostowania
def przyp_jeden():
    time.sleep(1)
    a=pyautogui.locateOnScreen('czy_chce_wycenic.png')
    b=pyautogui.center(a)
    pyautogui.click(b)
    time.sleep(1.5)

    try:

        if pyautogui.locateOnScreen('puste2.png'):
            #time.sleep(0.5)
            zd_status.append("Wyceniono")
            print('Wyceniono')
            print('0')

            
        elif pyautogui.locateOnScreen('brak_fak.png'):
            #time.sleep(0.5)
            a=pyautogui.locateOnScreen('ok.png')
            b=pyautogui.center(a)
            pyautogui.click(b)
            #time.sleep(0.5)
            zd_status.append("Brak możliwości wyceny")
            print('Brak możliwości wyceny')
            print('2')
            


        elif pyautogui.locateOnScreen('ok_1.png'):
            time.sleep(0.5)
            a=pyautogui.locateOnScreen('ok_1.png')
            b=pyautogui.center(a)
            pyautogui.click(b)
            #time.sleep(0.5)
            zd_status.append("Wyceniono")
            print('Zlecenie wyceniono')
            print('1')
           
        

        elif pyautogui.locateOnScreen('brak_ksiegowo.png'):
            while pyautogui.locateOnScreen('tak.png'):
                #time.sleep(0.5)
                a=pyautogui.locateOnScreen('tak.png')
                b=pyautogui.center(a)
                pyautogui.click(b)
                time.sleep(0.5)
            zd_status.append("Wyceniono")
            time.sleep(0.5)
            print('Zlecenie wyceniono')
            print('3')

            if pyautogui.locateOnScreen('ok_1.png'):
                #time.sleep(0.5)
                a=pyautogui.locateOnScreen('ok_1.png')
                b=pyautogui.center(a)
                pyautogui.click(b)
                #time.sleep(0.5)

    except Exception:
        print('Wystąpił błąd!')
        time.sleep(0.5)
        if pyautogui.locateOnScreen('ok_1.png'):
                #time.sleep(0.5)
                a=pyautogui.locateOnScreen('ok_1.png')
                b=pyautogui.center(a)
                pyautogui.click(b)
                #time.sleep(0.5)
                zd_status.append("Błąd")
    



#wycenianie
def cost():
    try:
              
        if pyautogui.locateOnScreen('wycen.png'):
            time.sleep(2)
            dolarek()
            przyp_jeden()
           # zd_status.append("Już było wycenione")
           # print('Już było wycenione')
           # time.sleep(2)
       # else:
        

            
        
       # elif pyautogui.locateOnScreen('wycenione.png'):
        #    zd_status.append("Już było wycenione")
         #   print('Już było wycenione')
          #  time.sleep(2)

    except Exception:
        print('Brak możliwości wyceny')
        zd_status.append("Brak możliwości wyceny")
        time.sleep(0.5)
        if pyautogui.locateOnScreen('ok_1.png'):
                #time.sleep(0.5)
                a=pyautogui.locateOnScreen('ok_1.png')
                b=pyautogui.center(a)
                pyautogui.click(b)
                #time.sleep(0.5)   
            

        

    
    
    
#________________________________________________________________________________________________________________
#open_browser(login, haslo) #uzupełnij

zd_list = []
zd_status = []

#otwiera przeglądarke i loguje
open_browser()

#pobiera liste zd
process_workbook('numeryzd.xlsx')
print(zd_list)
time.sleep(2)

#licz zlecenia
k=0


#wycenia
for zd in zd_list:
    okno()
    wklej(zd)
    print(f'Znaleziono zd :{zd}')
    zaznacz()
    #time.sleep(2)
    sprawdzenie()
    #cost()
    print('Procedura wyceniania przeprowadzona')
    kasuj()
    klik()
    k=k+1
    print (f'Zlecenie numer : {k}')
    print('')
    

#uzupełnia excela wyjściowego
excel_with_results('wycenione_statusy.xlsx')


